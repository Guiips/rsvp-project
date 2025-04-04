from typing import List
from fastapi import APIRouter, File, UploadFile, HTTPException, Depends, Request, Body, Form
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from models.evento import Evento, Convidado, StatusConvidado, EventoUpdate
from config.database import obter_db, get_database
from services.email_service import email_service
from bson import ObjectId
from datetime import datetime
import io
from openpyxl import load_workbook
from jose import jwt
from config.secrets import SECRET_KEY

# Cria o roteador de eventos
eventos_router = APIRouter()

# Templates config
templates = Jinja2Templates(directory="templates")

@eventos_router.post("/")
async def criar_evento(evento: Evento):
    """Cria um novo evento"""
    db = obter_db()
    try:
        evento_dict = evento.dict(exclude_unset=True)
        result = await db.eventos.insert_one(evento_dict)
        evento_dict['_id'] = str(result.inserted_id)
        
        print(f"Evento criado: {evento_dict}")
        return evento_dict
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@eventos_router.get("/")
async def listar_eventos():
    """Lista todos os eventos"""
    db = obter_db()
    eventos = []
    async for evento in db.eventos.find():
        if '_id' in evento:
            evento['_id'] = str(evento['_id'])
        if 'convidados' in evento:
            for convidado in evento.get('convidados', []):
                if '_id' in convidado:
                    convidado['_id'] = str(convidado['_id'])
        eventos.append(evento)
    return eventos

@eventos_router.get("/{evento_id}")
async def obter_evento(evento_id: str):
    """Obtém detalhes de um evento específico"""
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        evento = await db.eventos.find_one({"_id": object_id})
        if not evento:
            raise HTTPException(status_code=404, detail="Evento não encontrado")
        evento['_id'] = str(evento['_id'])
        return evento
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@eventos_router.delete("/{evento_id}")
async def excluir_evento(evento_id: str):
    """Exclui um evento"""
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        resultado = await db.eventos.delete_one({"_id": object_id})
        if resultado.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Evento não encontrado")
        return {"mensagem": "Evento deletado com sucesso"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@eventos_router.patch("/{evento_id}")
async def atualizar_evento(evento_id: str, evento_update: EventoUpdate):
    """Atualiza os dados de um evento"""
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        update_data = {k: v for k, v in evento_update.dict().items() if v is not None}
        
        resultado = await db.eventos.update_one(
            {"_id": object_id},
            {"$set": update_data}
        )
        
        if resultado.modified_count == 0:
            raise HTTPException(status_code=404, detail="Evento não encontrado")
        
        evento_atualizado = await db.eventos.find_one({"_id": object_id})
        evento_atualizado['_id'] = str(evento_atualizado['_id'])
        
        return evento_atualizado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@eventos_router.post("/{evento_id}/convidados")
async def adicionar_convidado(request: Request, evento_id: str, convidado: Convidado):
    """Adiciona um novo convidado ao evento"""
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        evento = await db.eventos.find_one({"_id": object_id})
        if not evento:
            raise HTTPException(status_code=404, detail="Evento não encontrado")

        # Gera os links de confirmação
        try:
            base_url = str(request.base_url).rstrip('/')
            link_confirmacao, link_recusa = email_service.gerar_tokens_para_evento(
                evento_id=evento_id,
                email_convidado=convidado.email,
                base_url=base_url
            )
        except Exception as token_error:
            print(f"Erro ao gerar tokens: {str(token_error)}")
            base_url = str(request.base_url).rstrip('/')
            link_confirmacao = f"{base_url}/api/eventos/confirmar-presenca/{evento_id}/{convidado.email}/sim"
            link_recusa = f"{base_url}/api/eventos/confirmar-presenca/{evento_id}/{convidado.email}/nao"

        # Adiciona o convidado
        convidado_dict = convidado.dict()
        await db.eventos.update_one(
            {"_id": object_id},
            {"$push": {"convidados": convidado_dict}}
        )

        # Envia o email de confirmação
        try:
            await email_service.enviar_email_confirmacao(
                email=convidado.email,
                nome=convidado.nome,
                evento_nome=evento['nome'],
                evento_data=evento['data'],
                evento_hora=evento['hora'],
                evento_local=evento['local'],
                link_confirmacao=link_confirmacao,
                link_recusa=link_recusa
            )
        except Exception as email_error:
            print(f"Erro ao enviar email: {str(email_error)}")

        evento_atualizado = await db.eventos.find_one({"_id": object_id})
        evento_atualizado['_id'] = str(evento_atualizado['_id'])
        return evento_atualizado

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@eventos_router.delete("/{evento_id}/convidados/excluir")
async def excluir_convidado(evento_id: str, dados: dict = Body(...)):
    """Exclui um convidado específico"""
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        resultado = await db.eventos.update_one(
            {"_id": object_id},
            {"$pull": {"convidados": {"email": dados['email']}}}
        )
        
        if resultado.modified_count == 0:
            raise HTTPException(status_code=404, detail="Convidado não encontrado")
        
        return {"mensagem": "Convidado excluído com sucesso"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@eventos_router.get("/confirmar-presenca/{evento_id}/{convidado_email}/{resposta}")
async def confirmar_presenca(request: Request, evento_id: str, convidado_email: str, resposta: str):
    """Processa a confirmação de presença do convidado"""
    try:
        evento_oid = ObjectId(evento_id)
        db = obter_db()

        evento = await db.eventos.find_one({"_id": evento_oid})
        if not evento:
            raise HTTPException(status_code=404, detail="Evento não encontrado")

        confirmado = resposta.lower() == 'sim'

        resultado = await db.eventos.update_one(
            {
                "_id": evento_oid,
                "convidados.email": convidado_email
            },
            {
                "$set": {
                    "convidados.$.confirmado": confirmado,
                    "convidados.$.data_confirmacao": datetime.utcnow(),
                    "convidados.$.status": (
                        StatusConvidado.CONFIRMADO if confirmado 
                        else StatusConvidado.RECUSADO
                    )
                }
            }
        )

        if resultado.modified_count == 0:
            raise HTTPException(status_code=404, detail="Convidado não encontrado")

        return templates.TemplateResponse(
            "confirmacao_sucesso.html",
            {
                "request": request,
                "confirmado": confirmado,
                "evento": evento,
                "email": convidado_email,
                "status": "confirmado" if confirmado else "recusado"
            }
        )

    except Exception as e:
        print(f"Erro ao processar confirmação: {str(e)}")
        return templates.TemplateResponse(
            "confirmacao_erro.html",
            {
                "request": request, 
                "mensagem": "Erro ao processar confirmação."
            }
        )

@eventos_router.post("/{evento_id}/convidados/salvar-observacoes/{email}")
async def salvar_observacoes(
    request: Request,
    evento_id: str,
    email: str,
    observacoes: str = Form(...)
):
    """Salva observações de um convidado que confirmou presença"""
    print(f"Tentando salvar observações para evento {evento_id} e email {email}")
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        
        resultado = await db.eventos.update_one(
            {
                "_id": object_id,
                "convidados.email": email
            },
            {
                "$set": {
                    "convidados.$.observacoes": observacoes,
                    "convidados.$.data_observacoes": datetime.utcnow()
                }
            }
        )
        
        if resultado.modified_count == 0:
            print("Convidado não encontrado ou não foi possível atualizar")
            return templates.TemplateResponse(
                "confirmacao_publica_agradecimento.html",
                {
                    "request": request,
                    "mensagem": "Não foi possível registrar sua resposta."
                }
            )
        
        print("Observações salvas com sucesso")
        return templates.TemplateResponse(
            "confirmacao_publica_agradecimento.html",
            {
                "request": request,
                "mensagem": "Suas observações foram salvas com sucesso!"
            }
        )
        
    except Exception as e:
        print(f"Erro ao salvar observações: {str(e)}")
        return templates.TemplateResponse(
            "confirmacao_publica_agradecimento.html",
            {
                "request": request,
                "mensagem": "Ocorreu um erro ao processar sua resposta."
            }
        )

@eventos_router.post("/{evento_id}/convidados/motivo-recusa/{email}")
async def salvar_motivo_recusa(request: Request, evento_id: str, email: str, motivo: str = Form(...)):
    """Salva o motivo da recusa de um convidado"""
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        
        resultado = await db.eventos.update_one(
            {
                "_id": object_id,
                "convidados.email": email
            },
            {
                "$set": {
                    "convidados.$.motivo_recusa": motivo,
                    "convidados.$.data_motivo": datetime.utcnow()
                }
            }
        )
        
        if resultado.modified_count == 0:
            return templates.TemplateResponse(
                "confirmacao_publica_agradecimento.html",
                {
                    "request": request,
                    "mensagem": "Não foi possível registrar sua resposta."
                }
            )
        
        return templates.TemplateResponse(
            "confirmacao_publica_agradecimento.html",
            {
                "request": request,
                "mensagem": "Agradecemos pelo seu feedback!"
            }
        )
        
    except Exception as e:
        print(f"Erro ao salvar motivo de recusa: {str(e)}")
        return templates.TemplateResponse(
            "confirmacao_publica_agradecimento.html",
            {
                "request": request,
                "mensagem": "Ocorreu um erro ao processar sua resposta."
            }
        )
        
        # [Mantendo todo o código anterior até a rota "salvar_motivo_recusa", e adicionando estas rotas adicionais:]

@eventos_router.post("/{evento_id}/convidados/importar")
async def importar_convidados(request: Request, evento_id: str, file: UploadFile = File(...)):
    """Importa lista de convidados de um arquivo Excel"""
    enviar_emails = False
    
    try:
        query_params = dict(request.query_params)
        enviar_emails_str = query_params.get('enviar_emails', 'false').lower()
        enviar_emails = enviar_emails_str == 'true'
    except:
        enviar_emails = False

    db = obter_db()
    try:
        if not file:
            raise HTTPException(status_code=400, detail="Nenhum arquivo enviado")

        contents = await file.read()
        
        if not contents:
            raise HTTPException(status_code=400, detail="Arquivo vazio")

        try:
            workbook = load_workbook(io.BytesIO(contents))
        except Exception as excel_error:
            print(f"Erro ao processar Excel: {excel_error}")
            raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo Excel: {str(excel_error)}")

        worksheet = workbook.active
        object_id = ObjectId(evento_id)
        evento = await db.eventos.find_one({"_id": object_id})
        
        if not evento:
            raise HTTPException(status_code=404, detail="Evento não encontrado")

        base_url = str(request.base_url).rstrip('/')
        convidados = []
        erros_email = []
        total_registros = 0
        registros_completos = 0
        registros_incompletos = 0

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            total_registros += 1
            nome = str(row[0]).strip() if row[0] and str(row[0]).strip() else "Sem nome"
            email = str(row[1]).strip() if len(row) > 1 and row[1] and str(row[1]).strip() else "Sem email"
            telefone = str(row[2]).strip() if len(row) > 2 and row[2] and str(row[2]).strip() else "Sem telefone"
            observacoes = str(row[3]).strip() if len(row) > 3 and row[3] and str(row[3]).strip() else "Sem observações"

            if nome != "Sem nome" and email != "Sem email" and '@' in email:
                registros_completos += 1
            else:
                registros_incompletos += 1

            convidado = {
                'nome': nome,
                'email': email,
                'telefone': telefone,
                'observacoes': observacoes,
                'status': StatusConvidado.PENDENTE
            }

            convidados.append(convidado)

            if enviar_emails and email != "Sem email" and '@' in email:
                try:
                    try:
                        link_confirmacao, link_recusa = email_service.gerar_tokens_para_evento(
                            evento_id=evento_id,
                            email_convidado=email,
                            base_url=base_url
                        )
                    except Exception as token_error:
                        print(f"Erro ao gerar tokens para {email}: {str(token_error)}")
                        link_confirmacao = f"{base_url}/api/eventos/confirmar-presenca/{evento_id}/{email}/sim"
                        link_recusa = f"{base_url}/api/eventos/confirmar-presenca/{evento_id}/{email}/nao"
                    
                    await email_service.enviar_email_confirmacao(
                        email=email,
                        nome=nome,
                        evento_nome=evento['nome'],
                        evento_data=evento['data'],
                        evento_hora=evento['hora'],
                        evento_local=evento['local'],
                        link_confirmacao=link_confirmacao,
                        link_recusa=link_recusa
                    )
                except Exception as email_error:
                    print(f"Erro ao enviar email para {email}: {str(email_error)}")
                    erros_email.append(email)

        if convidados:
            await db.eventos.update_one(
                {"_id": object_id},
                {"$push": {"convidados": {"$each": convidados}}}
            )

        return {
            "convidados": convidados,
            "total_importados": len(convidados),
            "total_registros": total_registros,
            "registros_completos": registros_completos,
            "registros_incompletos": registros_incompletos,
            "erros_email": erros_email,
            "emails_enviados": enviar_emails
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro não tratado na importação de convidados: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")

@eventos_router.post("/{evento_id}/convidados/enviar-emails")
async def enviar_emails_convidados(request: Request, evento_id: str):
    """Envia emails para todos os convidados do evento"""
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        evento = await db.eventos.find_one({"_id": object_id})
        
        if not evento:
            raise HTTPException(status_code=404, detail="Evento não encontrado")
        
        convidados = evento.get('convidados', [])
        
        if not convidados:
            return {"mensagem": "Não há convidados para enviar emails."}
        
        base_url = str(request.base_url).rstrip('/')
        emails_enviados = 0
        emails_com_erro = []
        
        for convidado in convidados:
            try:
                email = convidado.get('email')
                nome = convidado.get('nome')
                
                if not email or not nome or email == "Sem email" or '@' not in email:
                    continue
                
                try:
                    link_confirmacao, link_recusa = email_service.gerar_tokens_para_evento(
                        evento_id=evento_id,
                        email_convidado=email,
                        base_url=base_url
                    )
                except Exception as token_error:
                    print(f"Erro ao gerar tokens para {email}: {str(token_error)}")
                    link_confirmacao = f"{base_url}/api/eventos/confirmar-presenca/{evento_id}/{email}/sim"
                    link_recusa = f"{base_url}/api/eventos/confirmar-presenca/{evento_id}/{email}/nao"
                
                await email_service.enviar_email_confirmacao(
                    email=email,
                    nome=nome,
                    evento_nome=evento['nome'],
                    evento_data=evento['data'],
                    evento_hora=evento['hora'],
                    evento_local=evento['local'],
                    link_confirmacao=link_confirmacao,
                    link_recusa=link_recusa
                )
                
                emails_enviados += 1
                
            except Exception as email_error:
                print(f"Erro ao enviar email para {email}: {str(email_error)}")
                emails_com_erro.append(email)
        
        return {
            "mensagem": f"Emails enviados com sucesso: {emails_enviados}",
            "total_enviados": emails_enviados,
            "total_erros": len(emails_com_erro),
            "emails_com_erro": emails_com_erro
        }
        
    except Exception as e:
        print(f"Erro ao enviar emails: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@eventos_router.post("/{evento_id}/convidados/enviar-emails-template")
async def enviar_emails_com_template(request: Request, evento_id: str, dados: dict = Body(...)):
    """Envia emails para todos os convidados usando um template personalizado"""
    db = obter_db()
    try:
        object_id = ObjectId(evento_id)
        evento = await db.eventos.find_one({"_id": object_id})
        
        if not evento:
            raise HTTPException(status_code=404, detail="Evento não encontrado")
        
        convidados = evento.get('convidados', [])
        
        if not convidados:
            return {"mensagem": "Não há convidados para enviar emails."}
        
        assunto = dados.get('assunto', f"Convite para {evento['nome']}")
        corpo_template = dados.get('corpo', "")
        
        base_url = str(request.base_url).rstrip('/')
        emails_enviados = 0
        emails_com_erro = []
        
        for convidado in convidados:
            try:
                email = convidado.get('email')
                nome = convidado.get('nome')
                
                if not email or not nome or email == "Sem email" or '@' not in email:
                    continue
                
                try:
                    link_confirmacao, link_recusa = email_service.gerar_tokens_para_evento(
                        evento_id=evento_id,
                        email_convidado=email,
                        base_url=base_url
                    )
                except Exception as token_error:
                    print(f"Erro ao gerar tokens para {email}: {str(token_error)}")
                    link_confirmacao = f"{base_url}/api/eventos/confirmar-presenca/{evento_id}/{email}/sim"
                    link_recusa = f"{base_url}/api/eventos/confirmar-presenca/{evento_id}/{email}/nao"
                
                corpo_email = corpo_template
                corpo_email = corpo_email.replace("[Nome do Convidado]", nome)
                corpo_email = corpo_email.replace("[Link de Confirmação]", link_confirmacao)
                corpo_email = corpo_email.replace("[Link de Recusa]", link_recusa)
                
                try:
                    await email_service.enviar_email_html(
                        email=email,
                        assunto=assunto,
                        conteudo_html=corpo_email
                    )
                    
                    emails_enviados += 1
                except Exception as send_error:
                    print(f"Erro ao enviar email para {email}: {str(send_error)}")
                    emails_com_erro.append(email)
            except Exception as convidado_error:
                print(f"Erro ao processar convidado: {str(convidado_error)}")
                if 'email' in locals():
                    emails_com_erro.append(email)
        
        return {
            "mensagem": f"Emails enviados com sucesso: {emails_enviados}",
            "total_enviados": emails_enviados,
            "total_erros": len(emails_com_erro),
            "emails_com_erro": emails_com_erro
        }
        
    except Exception as e:
        print(f"Erro ao enviar emails: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@eventos_router.get("/confirmar/{token}")
async def confirmar_convite(token: str, request: Request):
    """Confirma a presença do convidado usando token JWT"""
    try:
        print(f"Token recebido para confirmação: {token}")
        
        try:
            db = obter_db()
            decoded_token = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            print(f"Token decodificado com sucesso: {decoded_token}")
            
            evento_id = decoded_token.get('evento_id')
            email_convidado = decoded_token.get('email')
            
            if not evento_id or not email_convidado:
                raise ValueError("Token não contém todas as informações necessárias")
            
            print(f"Processando confirmação para evento_id={evento_id}, email={email_convidado}")
            
            evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
            
            if not evento:
                return templates.TemplateResponse("confirmacao_erro.html", {
                    "request": request, 
                    "mensagem": "Evento não encontrado."
                })
            
            convidados = evento.get('convidados', [])
            convidado_atualizado = False
            convidado_info = None
            
            for convidado in convidados:
                if convidado.get('email') == email_convidado:
                    convidado['status'] = 'confirmado'
                    convidado_atualizado = True
                    convidado_info = convidado
                    break
            
            if convidado_atualizado:
                await db.eventos.update_one(
                    {"_id": ObjectId(evento_id)},
                    {"$set": {"convidados": convidados}}
                )
                
                print(f"Confirmação processada com sucesso. Status: confirmado")
                
                return templates.TemplateResponse("confirmacao_sucesso.html", {
                    "request": request, 
                    "evento": evento,
                    "status": "confirmado",
                    "email": email_convidado,
                    "convidado": convidado_info
                })
            else:
                return templates.TemplateResponse("confirmacao_erro.html", {
                    "request": request, 
                    "mensagem": "Convidado não encontrado."
                })
                
        except jwt.InvalidSignatureError:
            print("Erro: Assinatura do token inválida")
            return templates.TemplateResponse("confirmacao_erro.html", {
                "request": request, 
                "mensagem": "Link de confirmação inválido: assinatura inválida."
            })
        except jwt.DecodeError:
            print("Erro: Não foi possível decodificar o token")
            return templates.TemplateResponse("confirmacao_erro.html", {
                "request": request, 
                "mensagem": "Link de confirmação inválido: formato inválido."
            })
        except jwt.ExpiredSignatureError:
            print("Erro: Token expirado")
            return templates.TemplateResponse("confirmacao_erro.html", {
                "request": request, 
                "mensagem": "Link de confirmação expirado."
            })
        except Exception as token_error:
            print(f"Erro ao processar token: {str(token_error)}")
            return templates.TemplateResponse("confirmacao_erro.html", {
                "request": request, 
                "mensagem": f"Erro ao processar confirmação: {str(token_error)}"
            })
    
    except Exception as e:
       print(f"Erro ao processar confirmação: {str(e)}")
    return templates.TemplateResponse("confirmacao_erro.html", {
            "request": request, 
            "mensagem": f"Erro ao processar confirmação: {str(e)}"
        })

@eventos_router.get("/recusar/{token}")
async def recusar_convite(token: str, request: Request):
    """Recusa o convite usando token JWT"""
    try:
        print(f"Token recebido para recusa: {token}")
        
        try:
            db = obter_db()
            decoded_token = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            print(f"Token decodificado com sucesso: {decoded_token}")
            
            evento_id = decoded_token.get('evento_id')
            email_convidado = decoded_token.get('email')
            
            if not evento_id or not email_convidado:
                raise ValueError("Token não contém todas as informações necessárias")
            
            print(f"Processando recusa para evento_id={evento_id}, email={email_convidado}")
            
            evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
            
            if not evento:
                return templates.TemplateResponse("confirmacao_erro.html", {
                    "request": request, 
                    "mensagem": "Evento não encontrado."
                })
            
            convidados = evento.get('convidados', [])
            convidado_atualizado = False
            convidado_info = None
            
            for convidado in convidados:
                if convidado.get('email') == email_convidado:
                    convidado['status'] = 'recusado'
                    convidado_atualizado = True
                    convidado_info = convidado
                    break
            
            if convidado_atualizado:
                await db.eventos.update_one(
                    {"_id": ObjectId(evento_id)},
                    {"$set": {"convidados": convidados}}
                )
                
                print(f"Recusa processada com sucesso. Status: recusado")
                
                return templates.TemplateResponse("confirmacao_sucesso.html", {
                    "request": request, 
                    "evento": evento,
                    "status": "recusado",
                    "email": email_convidado,
                    "convidado": convidado_info
                })
            else:
                return templates.TemplateResponse("confirmacao_erro.html", {
                    "request": request, 
                    "mensagem": "Convidado não encontrado."
                })
                
        except jwt.InvalidSignatureError:
            print("Erro: Assinatura do token inválida")
            return templates.TemplateResponse("confirmacao_erro.html", {
                "request": request, 
                "mensagem": "Link de recusa inválido: assinatura inválida."
            })
        except jwt.DecodeError:
            print("Erro: Não foi possível decodificar o token")
            return templates.TemplateResponse("confirmacao_erro.html", {
                "request": request, 
                "mensagem": "Link de recusa inválido: formato inválido."
            })
        except jwt.ExpiredSignatureError:
            print("Erro: Token expirado")
            return templates.TemplateResponse("confirmacao_erro.html", {
                "request": request, 
                "mensagem": "Link de recusa expirado."
            })
        except Exception as token_error:
            print(f"Erro ao processar token: {str(token_error)}")
            return templates.TemplateResponse("confirmacao_erro.html", {
                "request": request, 
                "mensagem": f"Erro ao processar recusa: {str(token_error)}"
            })
    
    except Exception as e:
        print(f"Erro ao processar recusa: {str(e)}")
        return templates.TemplateResponse("confirmacao_erro.html", {
            "request": request, 
            "mensagem": f"Erro ao processar recusa: {str(e)}"
        })