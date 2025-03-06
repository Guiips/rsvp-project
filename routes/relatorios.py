from fastapi import APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import StreamingResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from routes.auth import get_current_user
from config.database import get_database
from pathlib import Path
from bson import ObjectId
import pandas as pd

router = APIRouter()

# Configuração dos templates
BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# Constantes de estilos
TITLE_FONT = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(name='Calibri', size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
STATUS_COLORS = {
    'confirmado': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'recusado': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'pendente': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
}


@router.get("/")
async def relatorios_index(request: Request):
    """Página inicial de relatórios"""
    # Carregar lista de eventos para o select
    db = get_database()
    eventos = []
    async for evento in db.eventos.find():
        if '_id' in evento:
            evento['_id'] = str(evento['_id'])
        eventos.append({
            "id": evento['_id'],
            "nome": evento.get('nome', 'Evento sem nome'),
            "data": evento.get('data', 'Sem data')
        })
    
    return templates.TemplateResponse(
        "relatorios/index.html",
        {
            "request": request, 
            "active_page": "relatorios", 
            "year": datetime.now().year,
            "eventos": eventos
        }
    )
    
@router.get("/eventos")
async def relatorio_eventos(request: Request):
    """Página de relatório de eventos"""
    # Carregar lista de eventos para o select
    db = get_database()
    eventos = []
    async for evento in db.eventos.find():
        if '_id' in evento:
            evento['_id'] = str(evento['_id'])
        eventos.append({
            "id": evento['_id'],
            "nome": evento.get('nome', 'Evento sem nome'),
            "data": evento.get('data', 'Sem data')
        })
    
    return templates.TemplateResponse(
        "relatorios/eventos.html",
        {
            "request": request, 
            "active_page": "relatorios_eventos", 
            "year": datetime.now().year,
            "eventos": eventos
        }
    )

@router.get("/convidados")
async def relatorio_convidados(request: Request):
    """Página de relatório de convidados"""
    return templates.TemplateResponse(
        "relatorios/convidados.html",
        {"request": request, "active_page": "relatorios_convidados", "year": datetime.now().year}
    )

@router.get("/gerar-relatorio")
async def gerar_relatorio():
    """Gera um relatório simples com todos os eventos e convidados"""
    db = get_database()
    
    # Consulta eventos e convidados
    eventos_convidados = await db.eventos.aggregate([
        {"$unwind": "$convidados"},
        {"$project": {
            "_id": 0,
            "evento_nome": "$nome",
            "evento_data": "$data",
            "nome": "$convidados.nome",
            "email": "$convidados.email",
            "telefone": "$convidados.telefone",
            "status": "$convidados.status"
        }}
    ]).to_list(length=None)
    
    # Preparar BytesIO para o arquivo Excel
    output = BytesIO()
    
    # Criar um workbook Excel usando openpyxl
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Lista de Convidados"
    
    # Adicionar título
    worksheet['A1'] = "RELATÓRIO GERAL DE CONVIDADOS"
    worksheet['A1'].font = TITLE_FONT
    worksheet['A1'].fill = TITLE_FILL
    worksheet.merge_cells('A1:F1')
    worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar cabeçalhos
    headers = ["Evento", "Data", "Nome", "E-mail", "Telefone", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Adicionar dados dos convidados
    for row_num, convidado in enumerate(eventos_convidados, 4):
        worksheet.cell(row=row_num, column=1).value = convidado.get('evento_nome', '')
        worksheet.cell(row=row_num, column=2).value = convidado.get('evento_data', '')
        worksheet.cell(row=row_num, column=3).value = convidado.get('nome', '')
        worksheet.cell(row=row_num, column=4).value = convidado.get('email', '')
        worksheet.cell(row=row_num, column=5).value = convidado.get('telefone', '')
        
        # Célula de status com formatação condicional
        status_cell = worksheet.cell(row=row_num, column=6)
        status_cell.value = convidado.get('status', '')
        status_cell.alignment = Alignment(horizontal='center')
        
        # Aplicar cor de fundo com base no status
        status = convidado.get('status', '')
        if status in STATUS_COLORS:
            status_cell.fill = STATUS_COLORS[status]
        
        # Adicionar bordas às células
        for col_num in range(1, 7):
            worksheet.cell(row=row_num, column=col_num).border = THIN_BORDER
    
    # Adicionar filtros nos cabeçalhos
    worksheet.auto_filter.ref = f"A3:F{row_num}" if eventos_convidados else "A3:F3"
    
    # Congelar os painéis para manter os cabeçalhos visíveis
    worksheet.freeze_panes = 'A4'
    
    # Ajustar largura das colunas automaticamente
    column_widths = {
        1: 30,  # Evento
        2: 15,  # Data
        3: 25,  # Nome
        4: 30,  # E-mail
        5: 15,  # Telefone
        6: 15,  # Status
    }
    
    for col_num, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Adicionar rodapé com data de geração
    footer_row = row_num + 2 if eventos_convidados else 5
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    worksheet.cell(row=footer_row, column=1).value = f"Relatório gerado em: {data_geracao}"
    worksheet.merge_cells(f'A{footer_row}:F{footer_row}')
    worksheet.cell(row=footer_row, column=1).font = Font(italic=True)
    
    # Adicionar contagem total
    total_row = footer_row - 1
    worksheet.cell(row=total_row, column=1).value = "Total de Convidados:"
    worksheet.cell(row=total_row, column=1).font = Font(bold=True)
    worksheet.cell(row=total_row, column=2).value = len(eventos_convidados)
    worksheet.cell(row=total_row, column=2).font = Font(bold=True)
    
    # Salvar o workbook
    workbook.save(output)
    output.seek(0)
    
    # Nome do arquivo com data atual
    data_atual = datetime.now().strftime("%Y-%m-%d")
    
    headers = {
        'Content-Disposition': f'attachment; filename="relatorio_geral_{data_atual}.xlsx"'
    }
    
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.get("/gerar-relatorio-evento/{evento_id}")
async def gerar_relatorio_evento(evento_id: str):
    """Gera um relatório para um evento específico com todas as informações em uma única planilha"""
    db = get_database()
    
    # Consulta informações do evento
    evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
    if not evento:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    # Consulta convidados do evento
    convidados_evento = await db.eventos.aggregate([
        {"$match": {"_id": ObjectId(evento_id)}},
        {"$unwind": "$convidados"},
        {"$project": {
            "_id": 0,
            "nome": "$convidados.nome",
            "email": "$convidados.email",
            "telefone": "$convidados.telefone",
            "status": "$convidados.status",
            "observacoes": "$convidados.observacoes"
        }}
    ]).to_list(length=None)
    
    # Preparar BytesIO para o arquivo Excel
    output = BytesIO()
    
    # Criar um workbook Excel usando openpyxl
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Relatório Completo"
    
    # Definir estilos
    title_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    
    section_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    subsection_font = Font(name='Calibri', size=12, bold=True)
    subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    
    label_font = Font(name='Calibri', size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== TÍTULO DO RELATÓRIO =====
    row = 1
    worksheet.merge_cells(f'A{row}:G{row}')
    cell = worksheet.cell(row=row, column=1)
    cell.value = "RELATÓRIO DETALHADO DE EVENTO"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtítulo com nome do evento
    row += 1
    worksheet.merge_cells(f'A{row}:G{row}')
    cell = worksheet.cell(row=row, column=1)
    cell.value = evento.get('nome', 'N/A')
    cell.font = Font(name='Calibri', size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    row += 2  # Espaço
    
    # ===== SEÇÃO 1: INFORMAÇÕES DO EVENTO =====
    worksheet.merge_cells(f'A{row}:G{row}')
    cell = worksheet.cell(row=row, column=1)
    cell.value = "INFORMAÇÕES DO EVENTO"
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal='center')
    
    # Informações básicas
    row += 2
    worksheet.merge_cells(f'A{row}:C{row}')
    cell = worksheet.cell(row=row, column=1)
    cell.value = "Dados Básicos"
    cell.font = subsection_font
    cell.fill = subsection_fill
    
    # Campos de informação
    info_fields = [
        {"label": "Nome do Evento:", "value": evento.get('nome', 'N/A')},
        {"label": "Data:", "value": evento.get('data', 'N/A')},
        {"label": "Horário:", "value": evento.get('hora', 'N/A')},
        {"label": "Local:", "value": evento.get('local', 'N/A')},
        {"label": "Responsável:", "value": evento.get('responsavel', 'N/A')},
        {"label": "Tipo de Evento:", "value": evento.get('categoria', 'N/A')},
        {"label": "Status:", "value": evento.get('status', 'Ativo')}
    ]
    
    for i, field in enumerate(info_fields):
        row += 1
        label_cell = worksheet.cell(row=row, column=1)
        label_cell.value = field["label"]
        label_cell.font = label_font
        label_cell.border = thin_border
        
        value_cell = worksheet.cell(row=row, column=2)
        value_cell.value = field["value"]
        value_cell.border = thin_border
        worksheet.merge_cells(f'B{row}:C{row}')
    
    # Descrição do evento
    row += 2
    worksheet.merge_cells(f'A{row}:C{row}')
    cell = worksheet.cell(row=row, column=1)
    cell.value = "Descrição do Evento"
    cell.font = subsection_font
    cell.fill = subsection_fill
    
    row += 1
    cell = worksheet.cell(row=row, column=1)
    cell.value = evento.get('descricao', 'Sem descrição disponível.')
    worksheet.merge_cells(f'A{row}:C{row+1}')
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = thin_border
    
    # ===== SEÇÃO 2: ESTATÍSTICAS DE PARTICIPAÇÃO (Lado direito) =====
    stats_start_row = 5  # Mesma linha das informações básicas
    worksheet.merge_cells(f'E{stats_start_row}:G{stats_start_row}')
    cell = worksheet.cell(row=stats_start_row, column=5)
    cell.value = "Estatísticas de Participação"
    cell.font = subsection_font
    cell.fill = subsection_fill
    
    # Calcular estatísticas
    total_convidados = len(convidados_evento)
    total_confirmados = len([c for c in convidados_evento if c.get('status') == 'confirmado'])
    total_recusados = len([c for c in convidados_evento if c.get('status') == 'recusado'])
    total_pendentes = len([c for c in convidados_evento if c.get('status') == 'pendente'])
    total_com_observacoes = len([c for c in convidados_evento if c.get('observacoes')])
    
    # Taxa de confirmação
    taxa_confirmacao = 0
    if total_convidados > 0:
        taxa_confirmacao = (total_confirmados / total_convidados) * 100
        
    # Classificar o sucesso do evento com base na taxa de confirmação
    classificacao = "Baixa"
    if taxa_confirmacao >= 75:
        classificacao = "Excelente"
        cor_classificacao = "00B050"  # Verde
    elif taxa_confirmacao >= 50:
        classificacao = "Boa"
        cor_classificacao = "92D050"  # Verde claro
    elif taxa_confirmacao >= 25:
        classificacao = "Regular"
        cor_classificacao = "FFEB9C"  # Amarelo
    else:
        classificacao = "Baixa"
        cor_classificacao = "FF7C80"  # Vermelho
    
    # Lista de estatísticas
    stats_fields = [
        {"label": "Total de Convidados:", "value": total_convidados},
        {"label": "Confirmados:", "value": f"{total_confirmados} ({(total_confirmados/total_convidados*100):.1f}%)" if total_convidados > 0 else "0 (0.0%)"},
        {"label": "Recusados:", "value": f"{total_recusados} ({(total_recusados/total_convidados*100):.1f}%)" if total_convidados > 0 else "0 (0.0%)"},
        {"label": "Pendentes:", "value": f"{total_pendentes} ({(total_pendentes/total_convidados*100):.1f}%)" if total_convidados > 0 else "0 (0.0%)"},
        {"label": "Com Observações:", "value": f"{total_com_observacoes} ({(total_com_observacoes/total_convidados*100):.1f}%)" if total_convidados > 0 else "0 (0.0%)"},
        {"label": "Taxa de Confirmação:", "value": f"{taxa_confirmacao:.1f}%"},
        {"label": "Classificação de Adesão:", "value": classificacao}
    ]
    
    for i, field in enumerate(stats_fields):
        stats_row = stats_start_row + i + 1
        label_cell = worksheet.cell(row=stats_row, column=5)
        label_cell.value = field["label"]
        label_cell.font = label_font
        label_cell.border = thin_border
        
        value_cell = worksheet.cell(row=stats_row, column=6)
        value_cell.value = field["value"]
        
        # Aplicar cor especial para a classificação
        if field["label"] == "Classificação de Adesão:":
            value_cell.font = Font(bold=True, color=cor_classificacao)
        
        value_cell.border = thin_border
        worksheet.merge_cells(f'F{stats_row}:G{stats_row}')
    
    # Pular para a próxima seção (após as informações e estatísticas)
    row = max(row + 3, stats_start_row + len(stats_fields) + 3)
    
    # ===== SEÇÃO 3: LISTA DE CONVIDADOS =====
    worksheet.merge_cells(f'A{row}:G{row}')
    cell = worksheet.cell(row=row, column=1)
    cell.value = "LISTA DE CONVIDADOS"
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal='center')
    
    row += 2
    
    # Legenda dos Status
    worksheet.merge_cells(f'A{row}:B{row}')
    cell = worksheet.cell(row=row, column=1)
    cell.value = "Legenda de Status:"
    cell.font = subsection_font
    
    # Cores de status
    status_colors = {
        'confirmado': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'recusado': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'pendente': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    }
    
    # Adicionar legenda
    row += 1
    for i, (status, fill) in enumerate(status_colors.items()):
        col = i * 2 + 1
        
        label_cell = worksheet.cell(row=row, column=col)
        label_cell.value = status.capitalize()
        label_cell.fill = fill
        label_cell.border = thin_border
        label_cell.alignment = Alignment(horizontal='center')
    
    row += 2
    
    # Cabeçalhos da tabela de convidados
    headers = ["Nome", "E-mail", "Telefone", "Status", "Observações"]
    for col_num, header in enumerate(headers, 1):
        col = col_num if col_num < 5 else col_num + 1  # Observações ocupa 2 colunas
        cell = worksheet.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        if col_num == 5:  # Observações ocupa 2 colunas
            worksheet.merge_cells(f'E{row}:F{row}')
    
    # Dados dos convidados
    for convidado_idx, convidado in enumerate(convidados_evento, 1):
        row += 1
        
        # Nome
        cell = worksheet.cell(row=row, column=1)
        cell.value = convidado.get('nome', '')
        cell.border = thin_border
        
        # E-mail
        cell = worksheet.cell(row=row, column=2)
        cell.value = convidado.get('email', '')
        cell.border = thin_border
        
        # Telefone
        cell = worksheet.cell(row=row, column=3)
        cell.value = convidado.get('telefone', '')
        cell.border = thin_border
        
        # Status com formatação condicional
        cell = worksheet.cell(row=row, column=4)
        cell.value = convidado.get('status', '')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        # Aplicar cor de fundo com base no status
        status = convidado.get('status', '')
        if status in status_colors:
            cell.fill = status_colors[status]
        
        # Observações (ocupa 2 colunas)
        cell = worksheet.cell(row=row, column=5)
        cell.value = convidado.get('observacoes', '')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border
        worksheet.merge_cells(f'E{row}:F{row}')
    
    # Se não houver convidados
    if not convidados_evento:
        row += 1
        cell = worksheet.cell(row=row, column=1)
        cell.value = "Nenhum convidado registrado para este evento."
        worksheet.merge_cells(f'A{row}:F{row}')
        cell.alignment = Alignment(horizontal='center')
    
    # ===== SEÇÃO 4: OBSERVAÇÕES IMPORTANTES (se houver) =====
    convidados_com_obs = [c for c in convidados_evento if c.get('observacoes')]
    
    if convidados_com_obs:
        row += 3
        
        worksheet.merge_cells(f'A{row}:G{row}')
        cell = worksheet.cell(row=row, column=1)
        cell.value = "OBSERVAÇÕES IMPORTANTES"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center')
        
        row += 2
        
        # Cabeçalhos da tabela de observações
        headers = ["Nome do Convidado", "Status", "Observações Detalhadas"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Observações ocupa várias colunas
            if col_num == 3:
                worksheet.merge_cells(f'C{row}:G{row}')
        
        # Dados das observações
        for obs_idx, convidado in enumerate(convidados_com_obs, 1):
            row += 1
            
            # Nome
            cell = worksheet.cell(row=row, column=1)
            cell.value = convidado.get('nome', '')
            cell.border = thin_border
            
            # Status com formatação condicional
            cell = worksheet.cell(row=row, column=2)
            cell.value = convidado.get('status', '')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Aplicar cor de fundo com base no status
            status = convidado.get('status', '')
            if status in status_colors:
                cell.fill = status_colors[status]
            
            # Observações detalhadas
            cell = worksheet.cell(row=row, column=3)
            cell.value = convidado.get('observacoes', '')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
            worksheet.merge_cells(f'C{row}:G{row}')
            
            # Ajustar altura para acomodar texto
            obs_text = convidado.get('observacoes', '')
            if obs_text:
                # Define diretamente a altura baseada no tamanho do texto
                # Evita o erro de comparação com NoneType
                linha_altura = min(100, max(20, len(obs_text) // 5))
                worksheet.row_dimensions[row].height = linha_altura
    
    # Adicionar rodapé com data de geração
    row += 3
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    cell = worksheet.cell(row=row, column=1)
    cell.value = f"Relatório gerado em: {data_geracao}"
    worksheet.merge_cells(f'A{row}:G{row}')
    cell.font = Font(italic=True)
    
    # Ajustar largura das colunas
    column_widths = {
        1: 25,  # Nome
        2: 30,  # E-mail
        3: 15,  # Telefone
        4: 12,  # Status
        5: 25,  # Observações começo
        6: 25,  # Observações continuação
        7: 10,  # Extra
    }
    
    for col_num, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Ajustar larguras automáticas para células com wrap text
    for row_num in range(1, row + 1):
        for cell in worksheet[row_num]:
            if cell.alignment and cell.alignment.wrap_text:
                # Definir uma altura fixa para células com texto envolvido
                # Isso evita problemas de comparação com NoneType
                worksheet.row_dimensions[row_num].height = 20
    
    # Salvar o workbook
    workbook.save(output)
    output.seek(0)
    
    # Nome do arquivo com nome do evento e data atual
    nome_evento_limpo = ''.join(e for e in evento.get('nome', '') if e.isalnum() or e in ' _-')
    data_atual = datetime.now().strftime("%Y-%m-%d")
    
    headers = {
        'Content-Disposition': f'attachment; filename="relatorio_{nome_evento_limpo}_{data_atual}.xlsx"'
    }
    
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
@router.get("/gerar-relatorio-completo")
async def gerar_relatorio_completo():
    """Gera um relatório completo com estatísticas e análises detalhadas de todos os eventos"""
    db = get_database()
    
    # Consulta todos os eventos
    eventos = await db.eventos.find().to_list(length=None)
    if not eventos:
        raise HTTPException(status_code=404, detail="Nenhum evento encontrado")
    
    # Consulta eventos e convidados
    eventos_convidados = await db.eventos.aggregate([
        {"$unwind": "$convidados"},
        {"$project": {
            "evento_id": "$_id",
            "evento_nome": "$nome",
            "evento_data": "$data",
            "evento_local": "$local",
            "nome": "$convidados.nome",
            "email": "$convidados.email",
            "telefone": "$convidados.telefone",
            "status": "$convidados.status",
            "observacoes": "$convidados.observacoes"
        }}
    ]).to_list(length=None)
    
    # Preparar BytesIO para o arquivo Excel
    output = BytesIO()
    
    # Criar um workbook Excel usando openpyxl
    workbook = openpyxl.Workbook()
    
    # Remover a planilha padrão
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # === Planilha de Resumo ===
    resume_sheet = workbook.create_sheet("Resumo Geral")
    
    # Adicionar título
    resume_sheet['A1'] = "RESUMO GERAL DE EVENTOS"
    resume_sheet['A1'].font = TITLE_FONT
    resume_sheet['A1'].fill = TITLE_FILL
    resume_sheet.merge_cells('A1:F1')
    resume_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar cabeçalhos da tabela de eventos
    headers = ["Nome do Evento", "Data", "Local", "Total de Convidados", "Confirmados", "Taxa de Confirmação"]
    for col_num, header in enumerate(headers, 1):
        cell = resume_sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Resumo por evento
    row_num = 4
    total_geral_convidados = 0
    total_geral_confirmados = 0
    
    for evento in eventos:
        evento_id = evento['_id']
        evento_convidados = [c for c in eventos_convidados if str(c.get('evento_id')) == str(evento_id)]
        
        total_convidados = len(evento_convidados)
        total_confirmados = len([c for c in evento_convidados if c.get('status') == 'confirmado'])
        
        # Calcular taxa de confirmação
        taxa_confirmacao = 0
        if total_convidados > 0:
            taxa_confirmacao = (total_confirmados / total_convidados) * 100
        
        # Adicionar linha de dados
        resume_sheet.cell(row=row_num, column=1).value = evento.get('nome', 'N/A')
        resume_sheet.cell(row=row_num, column=2).value = evento.get('data', 'N/A')
        resume_sheet.cell(row=row_num, column=3).value = evento.get('local', 'N/A')
        resume_sheet.cell(row=row_num, column=4).value = total_convidados
        resume_sheet.cell(row=row_num, column=5).value = total_confirmados
        resume_sheet.cell(row=row_num, column=6).value = f"{taxa_confirmacao:.1f}%"
        
        # Adicionar bordas
        for col_num in range(1, 7):
            resume_sheet.cell(row=row_num, column=col_num).border = THIN_BORDER
        
        # Incrementar contadores
        total_geral_convidados += total_convidados
        total_geral_confirmados += total_confirmados
        
        # Próxima linha
        row_num += 1
    
    # Linha de totais
    resume_sheet.cell(row=row_num, column=1).value = "TOTAL GERAL"
    resume_sheet.cell(row=row_num, column=1).font = SUBHEADER_FONT
    resume_sheet.merge_cells(f'A{row_num}:C{row_num}')
    
    resume_sheet.cell(row=row_num, column=4).value = total_geral_convidados
    resume_sheet.cell(row=row_num, column=4).font = SUBHEADER_FONT
    
    resume_sheet.cell(row=row_num, column=5).value = total_geral_confirmados
    resume_sheet.cell(row=row_num, column=5).font = SUBHEADER_FONT
    
    # Calcular taxa geral de confirmação
    taxa_geral = 0
    if total_geral_convidados > 0:
        taxa_geral = (total_geral_confirmados / total_geral_convidados) * 100
    
    resume_sheet.cell(row=row_num, column=6).value = f"{taxa_geral:.1f}%"
    resume_sheet.cell(row=row_num, column=6).font = SUBHEADER_FONT
    
    # Adicionar bordas à linha de totais
    for col_num in range(1, 7):
        resume_sheet.cell(row=row_num, column=col_num).border = THIN_BORDER
        resume_sheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Ajustar largura das colunas
    resume_sheet.column_dimensions['A'].width = 25
    resume_sheet.column_dimensions['B'].width = 15
    resume_sheet.column_dimensions['C'].width = 25
    resume_sheet.column_dimensions['D'].width = 20
    resume_sheet.column_dimensions['E'].width = 15
    resume_sheet.column_dimensions['F'].width = 20
    
    # Criar gráfico para total de convidados por evento
    if len(eventos) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Total de Convidados por Evento"
        chart.y_axis.title = "Número de Convidados"
        chart.x_axis.title = "Eventos"
        
        # Definir dados para o gráfico
        data = Reference(resume_sheet, min_col=4, min_row=3, max_row=row_num-1, max_col=4)
        cats = Reference(resume_sheet, min_col=1, min_row=4, max_row=row_num-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Adicionar o gráfico na planilha
        resume_sheet.add_chart(chart, "H4")
    
    # === Planilha de Estatísticas Detalhadas ===
    stats_sheet = workbook.create_sheet("Estatísticas Detalhadas")
    
    # Adicionar título
    stats_sheet['A1'] = "ESTATÍSTICAS DETALHADAS DE TODOS OS EVENTOS"
    stats_sheet['A1'].font = TITLE_FONT
    stats_sheet['A1'].fill = TITLE_FILL
    stats_sheet.merge_cells('A1:C1')
    stats_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar estatísticas gerais
    stats_sheet['A3'] = "Total de Eventos:"
    stats_sheet['B3'] = len(eventos)
    stats_sheet['A3'].font = SUBHEADER_FONT
    
    stats_sheet['A4'] = "Total de Convidados (Todos os Eventos):"
    stats_sheet['B4'] = total_geral_convidados
    stats_sheet['A4'].font = SUBHEADER_FONT
    
    stats_sheet['A5'] = "Total de Confirmações:"
    stats_sheet['B5'] = total_geral_confirmados
    stats_sheet['A5'].font = SUBHEADER_FONT
    
    taxa_confirmacao_geral = 0
    if total_geral_convidados > 0:
        taxa_confirmacao_geral = (total_geral_confirmados / total_geral_convidados) * 100
    
    stats_sheet['A6'] = "Taxa de Confirmação Geral:"
    stats_sheet['B6'] = f"{taxa_confirmacao_geral:.1f}%"
    stats_sheet['A6'].font = SUBHEADER_FONT
    
    # Adicionar estatísticas por status
    total_recusados = len([c for c in eventos_convidados if c.get('status') == 'recusado'])
    total_pendentes = len([c for c in eventos_convidados if c.get('status') == 'pendente'])
    
    stats_sheet['A8'] = "Distribuição por Status:"
    stats_sheet['A8'].font = SUBHEADER_FONT
    
    stats_sheet['A9'] = "Confirmados:"
    stats_sheet['B9'] = total_geral_confirmados
    if total_geral_convidados > 0:
        stats_sheet['C9'] = f"{(total_geral_confirmados/total_geral_convidados*100):.1f}%"
    else:
        stats_sheet['C9'] = "0.0%"
    
    stats_sheet['A10'] = "Recusados:"
    stats_sheet['B10'] = total_recusados
    if total_geral_convidados > 0:
        stats_sheet['C10'] = f"{(total_recusados/total_geral_convidados*100):.1f}%"
    else:
        stats_sheet['C10'] = "0.0%"
    
    stats_sheet['A11'] = "Pendentes:"
    stats_sheet['B11'] = total_pendentes
    if total_geral_convidados > 0:
        stats_sheet['C11'] = f"{(total_pendentes/total_geral_convidados*100):.1f}%"
    else:
        stats_sheet['C11'] = "0.0%"
    
    # Adicionar bordas
    for row in range(3, 12):
        for col in range(1, 4):
            if row == 7:
                continue
            cell = stats_sheet.cell(row=row, column=col)
            cell.border = THIN_BORDER
    
    # Criar gráfico de pizza para estatísticas por status
    if total_geral_convidados > 0:
        pie = PieChart()
        pie.title = "Distribuição por Status"
        labels = Reference(stats_sheet, min_col=1, min_row=9, max_row=11)
        data = Reference(stats_sheet, min_col=2, min_row=9, max_row=11)
        pie.add_data(data)
        pie.set_categories(labels)
        
        # Adicionar o gráfico na planilha
        stats_sheet.add_chart(pie, "E8")
    
    # Ajustar largura das colunas
    stats_sheet.column_dimensions['A'].width = 30
    stats_sheet.column_dimensions['B'].width = 15
    stats_sheet.column_dimensions['C'].width = 15
    
    # === Planilha de Dados Detalhados ===
    data_sheet = workbook.create_sheet("Dados Detalhados")
    
    # Adicionar título
    data_sheet['A1'] = "LISTA COMPLETA DE CONVIDADOS DE TODOS OS EVENTOS"
    data_sheet['A1'].font = TITLE_FONT
    data_sheet['A1'].fill = TITLE_FILL
    data_sheet.merge_cells('A1:G1')
    data_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar cabeçalhos
    headers = ["Evento", "Data", "Nome", "E-mail", "Telefone", "Status", "Observações"]
    for col_num, header in enumerate(headers, 1):
        cell = data_sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Adicionar dados dos convidados
    for row_num, convidado in enumerate(eventos_convidados, 4):
        data_sheet.cell(row=row_num, column=1).value = convidado.get('evento_nome', '')
        data_sheet.cell(row=row_num, column=2).value = convidado.get('evento_data', '')
        data_sheet.cell(row=row_num, column=3).value = convidado.get('nome', '')
        data_sheet.cell(row=row_num, column=4).value = convidado.get('email', '')
        data_sheet.cell(row=row_num, column=5).value = convidado.get('telefone', '')
        
        # Célula de status com formatação condicional
        status_cell = data_sheet.cell(row=row_num, column=6)
        status_cell.value = convidado.get('status', '')
        status_cell.alignment = Alignment(horizontal='center')
        
        # Aplicar cor de fundo com base no status
        status = convidado.get('status', '')
        if status in STATUS_COLORS:
            status_cell.fill = STATUS_COLORS[status]
        
        # Célula de observações
        data_sheet.cell(row=row_num, column=7).value = convidado.get('observacoes', '')
        
        # Adicionar bordas às células
        for col_num in range(1, 8):
            data_sheet.cell(row=row_num, column=col_num).border = THIN_BORDER
    
    # Adicionar filtros nos cabeçalhos
    data_sheet.auto_filter.ref = f"A3:G{row_num}" if eventos_convidados else "A3:G3"
    
    # Congelar os painéis para manter os cabeçalhos visíveis
    data_sheet.freeze_panes = 'A4'
    
    # Ajustar largura das colunas automaticamente
    column_widths = {
        1: 25,  # Evento
        2: 15,  # Data
        3: 25,  # Nome
        4: 30,  # E-mail
        5: 15,  # Telefone
        6: 15,  # Status
        7: 40,  # Observações
    }
    
    for col_num, width in column_widths.items():
        data_sheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Adicionar legenda de cores
    legend_row = row_num + 3 if eventos_convidados else 6
    
    data_sheet.cell(row=legend_row, column=1).value = "Legenda:"
    data_sheet.cell(row=legend_row, column=1).font = SUBHEADER_FONT
    
    # Confirmado
    data_sheet.cell(row=legend_row + 1, column=1).value = "Confirmado"
    data_sheet.cell(row=legend_row + 1, column=1).fill = STATUS_COLORS['confirmado']
    data_sheet.cell(row=legend_row + 1, column=1).border = THIN_BORDER
    
    # Recusado
    data_sheet.cell(row=legend_row + 2, column=1).value = "Recusado"
    data_sheet.cell(row=legend_row + 2, column=1).fill = STATUS_COLORS['recusado']
    data_sheet.cell(row=legend_row + 2, column=1).border = THIN_BORDER
    
    # Pendente
    data_sheet.cell(row=legend_row + 3, column=1).value = "Pendente"
    data_sheet.cell(row=legend_row + 3, column=1).fill = STATUS_COLORS['pendente']
    data_sheet.cell(row=legend_row + 3, column=1).border = THIN_BORDER
    
    # Adicionar rodapé com data de geração
    footer_row = legend_row + 5
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    data_sheet.cell(row=footer_row, column=1).value = f"Relatório gerado em: {data_geracao}"
    data_sheet.merge_cells(f'A{footer_row}:G{footer_row}')
    data_sheet.cell(row=footer_row, column=1).font = Font(italic=True)
    
    # Defina a planilha de resumo como ativa quando o arquivo for aberto
    workbook.active = resume_sheet
    
    # Salvar o workbook
    workbook.save(output)
    output.seek(0)
    
    # Nome do arquivo com data atual
    data_atual = datetime.now().strftime("%Y-%m-%d")
    
    headers = {
        'Content-Disposition': f'attachment; filename="relatorio_completo_{data_atual}.xlsx"'
    }
    
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )